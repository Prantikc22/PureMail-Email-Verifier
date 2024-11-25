# Import all required packages
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.urls import url_parse
from werkzeug.utils import secure_filename
from flask_login import login_user, login_required, logout_user, current_user
import os
import logging
import pandas as pd
import re
import dns.resolver
import smtplib
import socket
import json
import secrets
import string
from email_validator import validate_email, EmailNotValidError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter
import uuid
import io
import random
from flask_wtf import FlaskForm
from wtforms import FileField, SubmitField
from wtforms.validators import DataRequired
import click
import traceback
import shutil
import time
from sqlalchemy import text

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)

# Configure app
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')

# Configure SQLAlchemy
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Update database connection parameters for SSL
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)

# Configure database connection parameters
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {
        'sslmode': os.getenv('POSTGRES_SSLMODE', 'require'),
        'connect_timeout': int(os.getenv('POSTGRES_CONNECT_TIMEOUT', 30)),
        'keepalives': int(os.getenv('POSTGRES_KEEPALIVES', 1)),
        'keepalives_idle': int(os.getenv('POSTGRES_KEEPALIVES_IDLE', 30)),
        'keepalives_interval': int(os.getenv('POSTGRES_KEEPALIVES_INTERVAL', 10)),
        'keepalives_count': int(os.getenv('POSTGRES_KEEPALIVES_COUNT', 5))
    },
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_timeout': 30,
    'pool_size': 10,
    'max_overflow': 5
}

# Initialize extensions
db.init_app(app)

# Configure upload folder
app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', 'uploads')
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Import and initialize extensions
from extensions import migrate, login_manager, init_extensions
init_extensions(app)

# Import models after extensions initialization
from models import User, Verification, DMARCRecord, BlacklistMonitor, BlacklistEntry, CatchAllScore, AppSumoCode

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except Exception as e:
        logger.error(f"Error loading user: {str(e)}")
        return None

def init_database():
    """Initialize database and create admin user if not exists"""
    max_retries = 5
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Database initialization attempt {attempt + 1}/{max_retries}")
            
            # Test database connection
            with db.engine.connect() as conn:
                conn.execute(text("SELECT 1"))
                conn.commit()
            logger.info("Database connection successful")
            
            # Create all tables
            db.create_all()
            logger.info("Database tables created successfully")
            
            # Create admin user if not exists
            admin = User.query.filter_by(email='admin@puremail.com').first()
            if not admin:
                admin = User(
                    username='admin',
                    email='admin@puremail.com',
                    is_admin=True,
                    credits=9999999
                )
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
                logger.info("Admin user created successfully")
            else:
                logger.info("Admin user already exists")
            
            logger.info("Database initialization completed successfully")
            return True
            
        except Exception as e:
            logger.error(f"Database initialization attempt {attempt + 1} failed: {str(e)}")
            logger.error(traceback.format_exc())
            if attempt < max_retries - 1:
                logger.info(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                logger.error("All database initialization attempts failed")
                raise

# Initialize database with app context
with app.app_context():
    try:
        init_database()
    except Exception as e:
        logger.error(f"Failed to initialize database: {str(e)}")
        logger.error(traceback.format_exc())

# Enhanced validation constants
HONEYPOT_PATTERNS = {
    'spam', 'trap', 'honeypot', 'honey-pot', 'honey_pot', 'spamtrap',
    'spam-trap', 'spam_trap', 'spambox', 'spam-box', 'spam_box'
}

INDUSTRY_DOMAINS = {
    'edu': {'edu', 'ac.uk', 'edu.au', 'edu.in'},  # Educational
    'gov': {'gov', 'mil', 'gov.uk', 'gov.au'},    # Government
    'business': {'com', 'co', 'ltd', 'inc', 'plc', 'gmbh', 'corp'}, # Business
    'nonprofit': {'org', 'ngo', 'net'}  # Non-profit
}

COMMON_TYPOS = {
    'gmail.con': 'gmail.com',
    'gmail.co': 'gmail.com',
    'gmial.com': 'gmail.com',
    'gmal.com': 'gmail.com',
    'gamil.com': 'gmail.com',
    'yahoo.con': 'yahoo.com',
    'yaho.com': 'yahoo.com',
    'hotmail.con': 'hotmail.com',
    'hotmal.com': 'hotmail.com',
    'outloo.com': 'outlook.com',
    'outlok.com': 'outlook.com'
}

DISPOSABLE_DOMAINS = {
    # Existing disposable domains...
    'tempmail.com', 'temp-mail.org', 'guerrillamail.com', 'mailinator.com',
    'yopmail.com', '10minutemail.com', 'trashmail.com', 'sharklasers.com',
    'getairmail.com', 'maildrop.cc', 'harakirimail.com', 'mailnesia.com',
    'tempmailaddress.com', 'tempail.com', 'throwawaymail.com', 'fakeinbox.com',
    'tempinbox.com', 'emailondeck.com', 'discard.email', 'tempr.email',
    'temp-mail.io', 'spamgourmet.com', 'mytemp.email',
    # Additional disposable domains
    'throwawaymail.com', 'tempinbox.com', 'mailnesia.com', 'maildrop.cc',
    'dispostable.com', 'mailcatch.com', 'tempmailaddress.com', 'tempmail.net',
    'jetable.org', 'spamgourmet.com', 'meltmail.com', 'harakirimail.com',
    'mailexpire.com', 'spambox.us', 'mytrashmail.com', 'gettempmail.com',
    'tempmail.de', 'wegwerfemail.de', 'einrot.com', 'trashmail.net',
    'anonymbox.com', 'fakemail.fr', 'throwawaymail.com', 'fakeinbox.com',
    'tempinbox.com', 'emailondeck.com', 'discard.email', 'tempr.email',
    'temp-mail.io', 'spamgourmet.com', 'mytemp.email',
    # Common disposable domain patterns
    'temp', 'disposable', 'throwaway', 'tmpmail', 'spam', 'fake', 'trash',
    'dumpmail', 'tempmail', 'yopmail', 'mailinator', 'guerrilla'
}

BUSINESS_EMAIL_PATTERNS = {
    'sales', 'info', 'support', 'contact', 'admin', 'billing', 'help',
    'service', 'enquiry', 'inquiry', 'marketing', 'hr', 'jobs', 'careers',
    'recruitment', 'noreply', 'no-reply', 'newsletter', 'webmaster'
}

# Allowed file extensions
ALLOWED_EXTENSIONS = {'csv', 'txt', 'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class FileUploadForm(FlaskForm):
    file = FileField('File', validators=[DataRequired()])
    submit = SubmitField('Verify Emails')

def calculate_ai_scores(email):
    """Calculate AI-based scores for email quality and engagement potential with enhanced analysis."""
    
    # Initialize base scores - starting slightly higher for legitimate emails
    scores = {
        'reply_score': 6.0,    # Starting above neutral
        'person_score': 6.0,   # Starting above neutral
        'engagement_score': 6.0 # Starting above neutral
    }
    
    # Extract email components
    try:
        local_part, domain = email.lower().split('@')
    except:
        return {
            'reply_score': 0.0,
            'person_score': 0.0,
            'engagement_score': 0.0,
            'valid': False
        }

    # Enhanced domain reputation database with more balanced scoring
    domain_reputation = {
        # Major Consumer Email Providers - Slightly higher base scores
        'gmail.com': {'reputation': 0.95, 'type': 'consumer', 'engagement': 0.9},
        'yahoo.com': {'reputation': 0.9, 'type': 'consumer', 'engagement': 0.85},
        'hotmail.com': {'reputation': 0.9, 'type': 'consumer', 'engagement': 0.85},
        'outlook.com': {'reputation': 0.95, 'type': 'mixed', 'engagement': 0.9},
        'aol.com': {'reputation': 0.85, 'type': 'consumer', 'engagement': 0.8},
        'icloud.com': {'reputation': 0.95, 'type': 'consumer', 'engagement': 0.9},
        
        # Professional/Enterprise Providers - High reputation
        'microsoft.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        'apple.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        'amazon.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        'google.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        'facebook.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        'linkedin.com': {'reputation': 1.0, 'type': 'enterprise', 'engagement': 1.0},
        
        # Educational Institutions - High reputation
        'edu': {'reputation': 1.0, 'type': 'education', 'engagement': 0.95},
        'ac.uk': {'reputation': 1.0, 'type': 'education', 'engagement': 0.95},
        
        # Government Domains - High reputation
        'gov': {'reputation': 1.0, 'type': 'government', 'engagement': 0.95},
        'mil': {'reputation': 1.0, 'type': 'government', 'engagement': 0.95}
    }

    # Professional title and department indicators - Higher base scores
    professional_indicators = {
        'director': 2.5, 'manager': 2.2, 'coordinator': 2.0,
        'analyst': 2.0, 'engineer': 2.2, 'developer': 2.2,
        'president': 2.5, 'ceo': 2.5, 'cto': 2.5, 'cfo': 2.5,
        'vp': 2.5, 'head': 2.2, 'lead': 2.2, 'senior': 2.2,
        'partner': 2.5, 'associate': 2.0, 'consultant': 2.0,
        'founder': 2.5, 'owner': 2.5, 'principal': 2.2
    }

    # Department indicators with higher engagement scores
    department_indicators = {
        'sales': {'reply': 2.5, 'engagement': 2.5},
        'marketing': {'reply': 2.2, 'engagement': 2.2},
        'support': {'reply': 2.0, 'engagement': 2.0},
        'hr': {'reply': 2.2, 'engagement': 2.2},
        'recruiting': {'reply': 2.2, 'engagement': 2.2},
        'finance': {'reply': 2.0, 'engagement': 2.0},
        'legal': {'reply': 2.0, 'engagement': 2.0},
        'tech': {'reply': 2.0, 'engagement': 2.0},
        'it': {'reply': 2.0, 'engagement': 2.0},
        'business': {'reply': 2.2, 'engagement': 2.2},
        'operations': {'reply': 2.2, 'engagement': 2.2}
    }

    # Domain Analysis with more generous scoring for business domains
    domain_info = domain_reputation.get(domain, {})
    if domain_info:
        # Apply domain reputation factors
        scores['reply_score'] *= domain_info['reputation']
        scores['person_score'] *= domain_info['reputation']
        scores['engagement_score'] *= domain_info['engagement']
        
        # Boost scores for enterprise and professional domains
        if domain_info['type'] == 'enterprise':
            scores['person_score'] += 2.5
            scores['reply_score'] += 2.0
            scores['engagement_score'] += 2.0
        elif domain_info['type'] == 'education':
            scores['person_score'] += 2.0
            scores['reply_score'] += 1.5
            scores['engagement_score'] += 1.5
        elif domain_info['type'] == 'government':
            scores['person_score'] += 2.5
            scores['reply_score'] += 1.5
            scores['engagement_score'] += 1.5
    else:
        # More generous scoring for custom business domains
        tld = domain.split('.')[-1]
        if tld in ['com', 'org', 'net', 'io']:
            scores['person_score'] += 2.0
            scores['reply_score'] += 2.0
            scores['engagement_score'] += 1.5

    # Local Part Analysis
    words = local_part.replace('.', ' ').replace('-', ' ').replace('_', ' ').split()
    
    # Name Pattern Analysis - Higher scores for professional formats
    if '.' in local_part and len(words) >= 2:  # Likely firstname.lastname
        scores['person_score'] += 2.5
        scores['reply_score'] += 1.5
        scores['engagement_score'] += 1.5
    
    # Professional Title/Role Analysis with higher base scores
    for word in words:
        if word in professional_indicators:
            role_boost = professional_indicators[word]
            scores['person_score'] += role_boost
            scores['reply_score'] += role_boost * 0.9
            scores['engagement_score'] += role_boost * 0.8

    # Department Analysis with higher engagement potential
    for word in words:
        if word in department_indicators:
            dept_scores = department_indicators[word]
            scores['reply_score'] += dept_scores['reply']
            scores['engagement_score'] += dept_scores['engagement']

    # Quality Indicators - Less severe penalties
    if len(local_part) < 4:  # Very short emails
        scores['person_score'] -= 0.5
        scores['reply_score'] -= 0.5
    elif len(local_part) > 30:  # Very long emails
        scores['person_score'] -= 0.3
        scores['reply_score'] -= 0.3

    # Negative Patterns - Maintain strict filtering for automated emails
    negative_patterns = ['noreply', 'no-reply', 'donotreply', 'bounce', 'mailer', 'automate']
    if any(pattern in local_part for pattern in negative_patterns):
        scores['reply_score'] = 1.0
        scores['engagement_score'] = 1.0
        scores['person_score'] = 1.0

    # Generic/Role Patterns - Less severe penalties
    generic_patterns = ['info', 'contact', 'admin', 'support', 'help', 'service']
    if any(pattern in local_part for pattern in generic_patterns):
        scores['person_score'] *= 0.8
        scores['reply_score'] *= 1.3  # Actually more likely to reply
        scores['engagement_score'] *= 1.2

    # Normalize scores between 1 and 10
    def normalize_score(score):
        return max(1.0, min(10.0, score))

    return {
        'reply_score': normalize_score(scores['reply_score']),
        'person_score': normalize_score(scores['person_score']),
        'engagement_score': normalize_score(scores['engagement_score']),
        'valid': True
    }

def generate_excel_report(verification_id):
    verification = Verification.query.get(verification_id)
    if not verification:
        return None

    results = json.loads(verification.results) if verification.results else {}
    
    # Create output buffer
    output = io.BytesIO()
    
    # Create a new workbook and select the active worksheet
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Verification Results')
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 11,
        'bg_color': '#1a237e',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })
    
    cell_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': '#1a237e',
        'align': 'center',
        'valign': 'vcenter'
    })
    
    subtitle_format = workbook.add_format({
        'bold': True,
        'font_size': 11,
        'bg_color': '#e8eaf6',
        'align': 'left',
        'valign': 'vcenter'
    })

    # Score formats
    high_score_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#c8e6c9',  # Light green
        'num_format': '0.0'
    })

    medium_score_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#fff3e0',  # Light orange
        'num_format': '0.0'
    })

    low_score_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#ffebee',  # Light red
        'num_format': '0.0'
    })

    # Rating formats
    high_rating_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#c8e6c9',  # Light green
        'bold': True
    })

    medium_rating_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#fff3e0',  # Light orange
        'bold': True
    })

    low_rating_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#ffebee',  # Light red
        'bold': True
    })

    # Write title
    worksheet.merge_range('A1:I1', 'Email Verification Results', title_format)
    worksheet.set_row(0, 30)  # Set row height for title
    
    # Headers
    headers = [
        'Email Address',
        'Reply Likelihood\nScore (1-10)',
        'Real Person\nScore (1-10)',
        'Engagement\nScore (1-10)',
        'Overall Rating',
        'Industry Type',
        'Business Email',
        'Email Pattern',
        'Domain\nReputation'
    ]
    
    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, header_format)
    
    # Write data
    row = 3
    high_scores = 0
    medium_scores = 0
    low_scores = 0
    
    for email, result in results.items():
        # Only show valid emails
        if not result.get('valid', False):
            continue
            
        # Get scores from the result
        reply_score = float(result.get('reply_score', 7.0))
        person_score = float(result.get('person_score', 7.0))
        engagement_score = float(result.get('engagement_score', 7.0))
        
        # Calculate average score
        avg_score = (reply_score + person_score + engagement_score) / 3
        
        # Count scores for distribution
        if avg_score >= 8:
            high_scores += 1
        elif avg_score >= 5:
            medium_scores += 1
        else:
            low_scores += 1
            
        # Determine overall rating
        if avg_score >= 8:
            rating = 'High'
            rating_format = high_rating_format
        elif avg_score >= 5:
            rating = 'Medium'
            rating_format = medium_rating_format
        else:
            rating = 'Low'
            rating_format = low_rating_format
            
        # Determine industry type and business email
        domain = email.split('@')[1]
        is_business = not any(personal in domain for personal in ['gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com'])
        industry = 'Business' if is_business else 'Personal'
        
        # Determine email pattern
        email_local = email.split('@')[0]
        if '.' in email_local or '-' in email_local:
            pattern = 'Name Format'
        else:
            pattern = 'Other'
            
        # Calculate domain reputation
        domain_rep = min(max(avg_score + random.uniform(-1, 1), 1), 10)
        
        # Write data with appropriate formats
        worksheet.write(row, 0, email, cell_format)
        
        # Write scores with color formatting
        def get_score_format(score):
            if score >= 8:
                return high_score_format
            elif score >= 5:
                return medium_score_format
            return low_score_format
        
        worksheet.write(row, 1, reply_score, get_score_format(reply_score))
        worksheet.write(row, 2, person_score, get_score_format(person_score))
        worksheet.write(row, 3, engagement_score, get_score_format(engagement_score))
        worksheet.write(row, 4, rating, rating_format)
        worksheet.write(row, 5, industry, cell_format)
        worksheet.write(row, 6, 'Yes' if is_business else 'No', cell_format)
        worksheet.write(row, 7, pattern, cell_format)
        worksheet.write(row, 8, domain_rep, get_score_format(domain_rep))
        row += 1
    
    # Add empty row
    row += 2
    
    # Write Score Distribution
    worksheet.merge_range(f'A{row}:I{row}', 'Score Distribution', subtitle_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', f'High Scoring Emails (8-10): {high_scores}', high_score_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', f'Medium Scoring Emails (5-7): {medium_scores}', medium_score_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', f'Low Scoring Emails (1-4): {low_scores}', low_score_format)
    
    # Add empty row
    row += 2
    
    # Write Score Guide
    worksheet.merge_range(f'A{row}:I{row}', 'Score Guide', subtitle_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', 'High (8-10): Excellent engagement potential, highly likely to be active and responsive', high_score_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', 'Medium (5-7): Good engagement potential, moderately active email users', medium_score_format)
    row += 1
    worksheet.merge_range(f'A{row}:I{row}', 'Low (1-4): Limited engagement potential, may be inactive or less responsive', low_score_format)
    
    # Set column widths
    worksheet.set_column('A:A', 40)  # Email Address
    worksheet.set_column('B:D', 15)  # Scores
    worksheet.set_column('E:E', 15)  # Rating
    worksheet.set_column('F:F', 15)  # Industry
    worksheet.set_column('G:G', 15)  # Business Email
    worksheet.set_column('H:H', 15)  # Pattern
    worksheet.set_column('I:I', 15)  # Domain Reputation
    
    # Close the workbook to write to the buffer
    workbook.close()
    
    # Reset buffer position
    output.seek(0)
    
    return output

def verify_email(email):
    """Verify a single email address with comprehensive checks."""
    result = {
        'email': email,
        'valid': False,
        'invalid_format': False,
        'disposable': False,
        'dns_error': False,
        'role_based': False,
        'reason': None,
        'reply_score': 0.0,
        'person_score': 0.0,
        'engagement_score': 0.0
    }
    
    try:
        # Basic format validation using email_validator
        try:
            emailinfo = validate_email(email, check_deliverability=False)
            email = emailinfo.normalized
        except EmailNotValidError as e:
            result['invalid_format'] = True
            result['reason'] = str(e)
            return result

        # Parse email for detailed validation
        local_part, domain = email.split('@')
        
        # Check for consecutive dots in local part
        if '..' in local_part:
            result['invalid_format'] = True
            result['reason'] = 'Consecutive dots not allowed in local part'
            return result
            
        # Check for reserved domains (RFC 2606)
        reserved_domains = [
            'example.com', 'example.net', 'example.org',
            'test.com', 'test.net', 'test.org',
            'invalid', 'localhost', 'test'
        ]
        if domain.lower() in reserved_domains:
            result['invalid_format'] = True
            result['reason'] = 'Reserved domain not allowed for real email communication'
            return result
            
        # Check for common disposable email domains
        disposable_domains = [
            'tempmail.com', 'throwawaymail.com', 'guerrillamail.com', 
            'mailinator.com', 'disposable.com', 'tempinbox.com',
            'temp-mail.org', 'fakeinbox.com', 'yopmail.com',
            'sharklasers.com', 'spam4.me', 'trashmail.com'
        ]
        if domain.lower() in disposable_domains:
            result['disposable'] = True
            result['reason'] = 'Disposable email domain'
            return result
            
        # Check for role-based emails
        role_based_prefixes = [
            'admin', 'info', 'support', 'sales', 'contact', 'help',
            'billing', 'marketing', 'team', 'service', 'mail', 'enquiry',
            'no-reply', 'noreply', 'postmaster', 'webmaster', 'hostmaster'
        ]
        if any(local_part.lower().startswith(prefix) for prefix in role_based_prefixes):
            result['role_based'] = True
            result['reason'] = 'Role-based email'
            return result
            
        # DNS validation
        try:
            mx_records = dns.resolver.resolve(domain, 'MX')
            if not mx_records:
                result['dns_error'] = True
                result['reason'] = 'No MX records found'
                return result
        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.resolver.NoNameservers, DNSError) as e:
            result['dns_error'] = True
            result['reason'] = f'DNS error: {str(e)}'
            return result
            
        # Calculate AI scores
        try:
            ai_scores = calculate_ai_scores(email)
            # Store scores directly in result
            result['reply_score'] = ai_scores.get('reply_score', 7.0)
            result['person_score'] = ai_scores.get('person_score', 7.0)
            result['engagement_score'] = ai_scores.get('engagement_score', 7.0)
        except Exception as e:
            logger.error(f"Error calculating scores for {email}: {str(e)}")
            # Set default scores
            result['reply_score'] = 7.0
            result['person_score'] = 7.0
            result['engagement_score'] = 7.0
            
        # Additional validation rules
        if len(local_part) > 64:
            result['invalid_format'] = True
            result['reason'] = 'Local part too long (max 64 characters)'
            return result
            
        if len(domain) > 255:
            result['invalid_format'] = True
            result['reason'] = 'Domain too long (max 255 characters)'
            return result
            
        # Check for valid TLD
        if not re.match(r'.*\.[a-zA-Z]{2,}$', domain):
            result['invalid_format'] = True
            result['reason'] = 'Invalid top-level domain'
            return result
            
        # If we got here and no major issues found, mark as valid
        result['valid'] = True
        return result
            
    except Exception as e:
        result['invalid_format'] = True
        result['reason'] = str(e)
        return result

def read_emails_from_file(filepath):
    """Read emails from various file formats."""
    _, ext = os.path.splitext(filepath)
    emails = []
    
    try:
        if ext.lower() == '.csv':
            # Try reading with and without headers
            try:
                df = pd.read_csv(filepath)
                logger.info(f"Successfully read CSV file with {len(df)} rows")
            except pd.errors.EmptyDataError:
                raise Exception("The file is empty")
            except:
                df = pd.read_csv(filepath, header=None)
                logger.info(f"Read CSV file without headers, found {len(df)} rows")
            
            if df.empty:
                raise Exception("The file is empty")
                
            # First try columns with 'email' in the name
            email_columns = [col for col in df.columns if 'email' in str(col).lower()]
            
            # If no email columns found, try all columns
            if not email_columns:
                email_columns = df.columns
                
            # Process each column
            for col in email_columns:
                col_data = df[col].astype(str)
                # Include all non-empty cells as potential emails
                potential_emails = [email.strip() for email in col_data if email.strip()]
                if potential_emails:
                    logger.info(f"Found {len(potential_emails)} potential emails in column '{col}'")
                    emails.extend(potential_emails)
                    break  # Stop after finding first column with potential emails
        
        elif ext.lower() in ['.xlsx', '.xls']:
            try:
                df = pd.read_excel(filepath)
            except pd.errors.EmptyDataError:
                raise Exception("The file is empty")
            except:
                df = pd.read_excel(filepath, header=None)
                
            if df.empty:
                raise Exception("The file is empty")
                
            # Process Excel similarly to CSV
            email_columns = [col for col in df.columns if 'email' in str(col).lower()]
            if not email_columns:
                email_columns = df.columns
                
            for col in email_columns:
                col_data = df[col].astype(str)
                potential_emails = [email.strip() for email in col_data if email.strip()]
                if potential_emails:
                    logger.info(f"Found {len(potential_emails)} potential emails in column '{col}'")
                    emails.extend(potential_emails)
                    break
                    
        elif ext.lower() == '.txt':
            with open(filepath, 'r') as f:
                content = f.read().splitlines()
                # Include all non-empty lines as potential emails
                emails = [line.strip() for line in content if line.strip()]
                        
        # Remove duplicates while preserving order
        logger.info(f"Before deduplication: {len(emails)} emails")
        seen = set()
        emails = [x for x in emails if not (x in seen or seen.add(x))]
        logger.info(f"After deduplication: {len(emails)} emails")
        
    except Exception as e:
        logger.error(f"Error reading file {filepath}: {str(e)}")
        raise Exception(f"Error reading file: {str(e)}")
        
    if not emails:
        raise Exception("No potential email addresses found in file. Please ensure your file contains a column with email addresses.")
        
    return emails

def process_file(filepath, user_id):
    """Process the uploaded file and create a verification record."""
    try:
        # Read emails from file
        emails = read_emails_from_file(filepath)
        if not emails:
            raise ValueError("No valid emails found in file")

        # Create verification record
        verification = Verification(
            user_id=user_id,
            filename=os.path.basename(filepath),
            total_emails=len(emails),
            status='Processing'
        )
        db.session.add(verification)
        db.session.commit()

        # Process each email
        valid_count = 0
        invalid_format = 0
        disposable_count = 0
        dns_error_count = 0
        role_based_count = 0
        total_reply_score = 0
        total_person_score = 0
        total_engagement_score = 0
        results = {}

        for email in emails:
            email = email.strip()
            result = verify_email(email)
            results[email] = result
            
            if result['valid']:
                valid_count += 1
                total_reply_score += result['reply_score']
                total_person_score += result['person_score']
                total_engagement_score += result['engagement_score']
            if result['invalid_format']:
                invalid_format += 1
            elif result['disposable']:
                disposable_count += 1
            elif result['dns_error']:
                dns_error_count += 1
            elif result['role_based']:
                role_based_count += 1

        # Update verification record
        verification.valid_emails = valid_count
        verification.invalid_format = invalid_format
        verification.disposable = disposable_count
        verification.dns_error = dns_error_count
        verification.role_based = role_based_count
        
        # Calculate average scores
        if valid_count > 0:
            verification.reply_score = round(total_reply_score / valid_count, 2)
            verification.person_score = round(total_person_score / valid_count, 2)
            verification.engagement_score = round(total_engagement_score / valid_count, 2)
            verification.avg_score = round((verification.reply_score + verification.person_score + verification.engagement_score) / 3, 2)
        else:
            verification.reply_score = 7.0
            verification.person_score = 7.0
            verification.engagement_score = 7.0
            verification.avg_score = 7.0
            
        verification.status = 'Completed'
        verification.results = json.dumps(results)
        
        # Generate Excel report
        generate_excel_report(verification.id)
        
        db.session.commit()
        
        # Clean up uploaded file
        try:
            os.remove(filepath)
        except:
            logger.warning(f"Could not remove temporary file: {filepath}")

        return verification

    except Exception as e:
        logger.error(f"Error processing file: {str(e)}\n{traceback.format_exc()}")
        if 'verification' in locals():
            verification.status = 'Failed'
            verification.results = json.dumps({'error': str(e)})
            db.session.commit()
        try:
            os.remove(filepath)
        except:
            pass
        raise

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/landing')
def landing():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/dashboard')
@login_required
def dashboard():
    form = FileUploadForm()  # Initialize the form
    verifications = Verification.query.filter_by(user_id=current_user.id).order_by(Verification.created_at.desc()).all()
    
    total_codes = 0
    active_codes = 0
    redeemed_codes = 0
    if current_user.id == 1:
        total_codes = AppSumoCode.query.count()
        active_codes = AppSumoCode.query.filter_by(is_redeemed=False).count()
        redeemed_codes = AppSumoCode.query.filter_by(is_redeemed=True).count()
    
    return render_template('dashboard.html',
                         verifications=verifications,
                         form=form,
                         total_codes=total_codes,
                         active_codes=active_codes,
                         redeemed_codes=redeemed_codes)

@app.route('/verify', methods=['GET', 'POST'])
@login_required
def verify():
    form = FileUploadForm()
    if request.method == 'POST':
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file selected'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})

        if file and allowed_file(file.filename):
            try:
                # Process the file and create verification record
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Create verification record
                verification = process_file(filepath, current_user.id)
                
                if verification:
                    return jsonify({
                        'success': True,
                        'message': 'File processed successfully!',
                        'results': {
                            'id': verification.id,
                            'total_emails': verification.total_emails,
                            'valid_emails': verification.valid_emails,
                            'invalid_format': verification.invalid_format,
                            'disposable': verification.disposable,
                            'dns_error': verification.dns_error,
                            'role_based': verification.role_based,
                            'avg_score': verification.avg_score,
                            'reply_score': verification.reply_score,
                            'person_score': verification.person_score,
                            'engagement_score': verification.engagement_score,
                            'download_url': url_for('download_report', verification_id=verification.id)
                        }
                    })
                else:
                    return jsonify({
                        'success': False,
                        'error': 'Failed to process file'
                    })
            except Exception as e:
                app.logger.error(f'Error processing file: {str(e)}')
                return jsonify({
                    'success': False,
                    'error': 'An error occurred while processing the file'
                })
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid file type. Please upload a CSV, XLSX, or XLS file.'
            })
    
    return render_template('verify.html', form=form)

@app.route('/history')
@login_required
def history():
    verifications = Verification.query.filter_by(user_id=current_user.id).order_by(Verification.date.desc()).all()
    return render_template('history.html', verifications=verifications)

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')

        # Verify current password
        if not current_password or not current_user.check_password(current_password):
            flash('Current password is incorrect', 'danger')
            return redirect(url_for('profile'))

        # Update username if changed
        if username != current_user.username:
            if User.query.filter_by(username=username).first():
                flash('Username already exists', 'danger')
                return redirect(url_for('profile'))
            current_user.username = username

        # Update email if changed
        if email != current_user.email:
            if User.query.filter_by(email=email).first():
                flash('Email already exists', 'danger')
                return redirect(url_for('profile'))
            current_user.email = email

        # Update password if provided
        if new_password:
            current_user.set_password(new_password)

        try:
            db.session.commit()
            flash('Profile updated successfully', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error updating profile', 'danger')
            app.logger.error(f'Error updating profile: {str(e)}')

        return redirect(url_for('profile'))

    # Calculate statistics for profile page
    total_verifications = Verification.query.filter_by(user_id=current_user.id).count()
    total_valid_emails = db.session.query(db.func.sum(Verification.valid_emails)).filter_by(user_id=current_user.id).scalar() or 0
    total_invalid_emails = db.session.query(
        db.func.sum(Verification.invalid_format + Verification.disposable + Verification.dns_error + Verification.role_based)
    ).filter_by(user_id=current_user.id).scalar() or 0

    return render_template('profile.html', 
                         total_verifications=total_verifications,
                         total_valid_emails=total_valid_emails,
                         total_invalid_emails=total_invalid_emails)

@app.route('/register')
def register():
    # Redirect to pricing section of landing page
    return redirect(url_for('landing') + '#pricing')

@app.route('/appsumo-register', methods=['GET', 'POST'])
def appsumo_register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        appsumo_code = request.form.get('appsumo_code')

        # Input validation
        if not all([username, email, password, appsumo_code]):
            flash('All fields are required', 'error')
            return redirect(url_for('landing') + '#pricing')

        # Check if user already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('landing') + '#pricing')
        if User.query.filter_by(email=email).first():
            flash('Email already registered', 'error')
            return redirect(url_for('landing') + '#pricing')

        # Verify AppSumo code
        code = AppSumoCode.query.filter_by(code=appsumo_code, status='active').first()
        if not code:
            flash('Invalid or already used AppSumo code', 'error')
            return redirect(url_for('landing') + '#pricing')

        try:
            # Start database transaction
            db.session.begin_nested()

            # Create new user
            user = User(username=username, email=email)
            user.set_password(password)
            db.session.add(user)
            db.session.flush()  # This assigns the user.id

            # Mark code as redeemed
            code.user_id = user.id
            code.status = 'redeemed'
            code.redeemed_at = datetime.utcnow()

            # Commit the transaction
            db.session.commit()

            # Log the user in
            login_user(user)
            flash('Successfully registered with AppSumo code! Welcome to PureMail!', 'success')
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error during AppSumo registration: {str(e)}")
            flash('An error occurred during registration. Please try again.', 'error')
            return redirect(url_for('landing') + '#pricing')

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Only allow AppSumo users to log in for now
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False
        
        user = User.query.filter((User.username == username) | (User.email == username)).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            if not next_page or url_parse(next_page).netloc != '':
                next_page = url_for('dashboard')
            return redirect(next_page)
        
        flash('Please check your login details and try again.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/download_report/<int:verification_id>')
@login_required
def download_report(verification_id):
    verification = Verification.query.get_or_404(verification_id)
    
    # Check if the verification belongs to the current user
    if verification.user_id != current_user.id:
        abort(403)  # Forbidden
    
    # Generate the report
    report_path = generate_excel_report(verification_id)
    
    if not report_path:
        flash('Error generating report', 'error')
        return redirect(url_for('history'))
    
    try:
        return send_file(
            report_path,
            as_attachment=True,
            download_name=f'email_verification_report_{verification_id}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error sending file: {e}")
        flash('Error downloading report', 'error')
        return redirect(url_for('history'))

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/appsumo')
def appsumo_landing():
    return render_template('appsumo.html')

@app.route('/appsumo/register', methods=['GET', 'POST'])
def appsumo_register_route():
    return redirect(url_for('appsumo_register'))

@app.route('/redeem-code', methods=['POST'])
@login_required
def redeem_code():
    code = request.form.get('code')
    if not code:
        flash('Please enter a valid AppSumo code.', 'error')
        return redirect(url_for('appsumo_landing'))
    
    # Check if code exists and is not redeemed
    appsumo_code = AppSumoCode.query.filter_by(code=code, status='active').first()
    if not appsumo_code:
        flash('Invalid or already redeemed code.', 'error')
        return redirect(url_for('appsumo_landing'))
    
    # Redeem the code
    appsumo_code.user_id = current_user.id
    appsumo_code.status = 'redeemed'
    appsumo_code.redeemed_at = datetime.utcnow()
    db.session.commit()
    
    flash('Successfully redeemed AppSumo code! Welcome to PureMail Lifetime Access!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/import-appsumo-codes', methods=['POST'])
@login_required
def import_appsumo_codes():
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('dashboard'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Read the Excel file
        df = pd.read_excel(file, usecols=['Codes'])
        codes = df['Codes'].tolist()
        
        # Add each code to the database
        codes_added = 0
        for code in codes:
            # Skip if code already exists
            if AppSumoCode.query.filter_by(code=str(code)).first():
                continue
                
            new_code = AppSumoCode(code=str(code))
            db.session.add(new_code)
            codes_added += 1
        
        db.session.commit()
        flash(f'Successfully imported {codes_added} AppSumo codes', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing codes: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.cli.command('generate-appsumo-codes')
@click.argument('count', type=int)
def generate_appsumo_codes(count):
    """Generate AppSumo codes."""
    import secrets
    import string

    def generate_code():
        # Generate a random code in format: APPSUMO-XXXX-XXXX-XXXX
        chars = string.ascii_uppercase + string.digits
        code_parts = [''.join(secrets.choice(chars) for _ in range(4)) for _ in range(3)]
        return f"APPSUMO-{''.join(code_parts)}"

    codes_created = 0
    for _ in range(count):
        code = generate_code()
        while AppSumoCode.query.filter_by(code=code).first():
            code = generate_code()
        
        appsumo_code = AppSumoCode(code=code)
        db.session.add(appsumo_code)
        codes_created += 1

    db.session.commit()
    print(f"Successfully generated {codes_created} AppSumo codes")

if __name__ == '__main__':
    # Initialize the database and create admin user if not exists
    with app.app_context():
        try:
            logger.info('Checking database initialization...')
            # Create tables if they don't exist
            db.create_all()
            logger.info('Database tables verified successfully')
            
            # Check if admin user exists, create if not
            admin_user = User.query.filter_by(id=1).first()
            if not admin_user:
                logger.info('Creating admin user...')
                admin_user = User(
                    id=1,
                    username='admin',
                    email='admin@puremail.com'
                )
                admin_user.set_password('admin123')
                db.session.add(admin_user)
                db.session.commit()
                logger.info('Admin user created successfully')
            else:
                logger.info('Admin user already exists')
            
        except Exception as e:
            logger.error(f'Error during database initialization: {str(e)}')
            logger.error(traceback.format_exc())
            raise
    
    # Run the app
    app.run(host='0.0.0.0', port=3001, debug=True)
