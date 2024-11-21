"""
Excel report generation module for PureMail
"""
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_excel_report(results: List[Dict], stats: Dict, filepath: str) -> None:
    """Create a detailed Excel report with validation results."""
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add title and timestamp
    ws_summary['A1'] = "Email Verification Report"
    ws_summary['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Style the header
    header_font = Font(bold=True, size=14)
    ws_summary['A1'].font = header_font
    
    # Add statistics
    ws_summary['A4'] = "Verification Statistics"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    row = 5
    for key, value in stats.items():
        ws_summary[f'A{row}'] = key.replace('_', ' ').title()
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Detailed results sheet
    ws_details = wb.create_sheet("Detailed Results")
    
    # Headers for detailed results
    headers = [
        'Email',
        'Valid',
        'Is Business',
        'Is Suspicious',
        'Possible Typo',
        'Suggested Correction',
        'Is Disposable',
        'High Risk TLD',
        'Spam Score',
        'Security Warnings',
        'Suggestions'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add results
    row = 2
    for result in results:
        pattern_analysis = result.get('pattern_analysis', {})
        security = result.get('security', {})
        
        # Write data
        data = [
            result['email'],
            result['is_valid'],
            pattern_analysis.get('is_business', False),
            pattern_analysis.get('is_suspicious', False),
            bool(pattern_analysis.get('possible_typo')),
            pattern_analysis.get('possible_typo', ''),
            security.get('is_disposable', False),
            security.get('is_high_risk_tld', False),
            security.get('spam_score', 0),
            '\n'.join(security.get('security_warnings', [])),
            '\n'.join(pattern_analysis.get('suggestions', []))
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_details.cell(row=row, column=col, value=value)
            
            # Color invalid emails
            if col == 1 and not result['is_valid']:
                for c in range(1, len(headers) + 1):
                    ws_details.cell(row=row, column=c).fill = PatternFill(
                        start_color="FFE6E6",
                        end_color="FFE6E6",
                        fill_type="solid"
                    )
        
        row += 1
    
    # Auto-adjust column widths
    for sheet in [ws_summary, ws_details]:
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filepath)
